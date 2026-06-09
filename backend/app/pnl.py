from __future__ import annotations
import datetime as dt
import math
import os
import time
from pathlib import Path
from typing import Any


def _mock_enabled() -> bool:
    """손익현황(2번째 탭) Mock 모드 스위치. 기본 ON(데모용).

    기존 엑셀 기반 로직은 그대로 보존하며, 이 플래그가 켜졌을 때만
    mock_portfolio 모듈의 데이터를 반환한다. 끄려면 ATLAS_PNL_MOCK=0.
    """
    return os.environ.get("ATLAS_PNL_MOCK", "1").strip().lower() not in ("0", "false", "off", "no", "")

DATA_DIR   = Path(__file__).resolve().parents[1] / "data"
EXCEL_PATH = DATA_DIR / "주간주식시황_V10_마스터파일.xlsx"


# ---------------------------------------------------------------------------
# Module-level sheet-row cache (Task A)
# ---------------------------------------------------------------------------

_CACHE_TTL = 60  # seconds

# Cache entry: {"mtime": float, "loaded_at": float, "sheets": dict[str, list[tuple]]}
_sheets_cache: dict[str, Any] | None = None


def _load_sheets_cached() -> dict[str, list]:
    """Return dict of sheet_name -> list-of-row-tuples (values_only), cached by mtime+TTL.

    SAFER than caching the workbook object: we parse everything once, close the
    file, and store plain Python tuples.  Subsequent callers get the same lists
    without any open file handles.
    """
    global _sheets_cache
    import openpyxl

    if not EXCEL_PATH.exists():
        return {}

    try:
        mtime = EXCEL_PATH.stat().st_mtime
    except OSError:
        return {}

    now = time.monotonic()
    if (
        _sheets_cache is not None
        and _sheets_cache["mtime"] == mtime
        and (now - _sheets_cache["loaded_at"]) < _CACHE_TTL
    ):
        return _sheets_cache["sheets"]

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        sheets: dict[str, list] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return _sheets_cache["sheets"] if _sheets_cache else {}

    _sheets_cache = {"mtime": mtime, "loaded_at": now, "sheets": sheets}
    return sheets


def _get_sheet_rows(name_fragment: str) -> list:
    """Return rows for the first sheet whose name contains name_fragment, or []."""
    sheets = _load_sheets_cached()
    for sname, rows in sheets.items():
        if name_fragment in sname:
            return rows
    return []


# ---------------------------------------------------------------------------
# Excel reader helpers
# ---------------------------------------------------------------------------

def _load_wb():
    """Open workbook directly (used by functions that need the wb object for
    _sheet_to_dicts/asset_master/etc.).  Callers are responsible for wb.close()."""
    import openpyxl
    if not EXCEL_PATH.exists():
        return None
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def _sheet_to_dicts(wb, name_fragment: str) -> list[dict]:
    """시트명에 name_fragment 포함된 시트 → 헤더-값 딕셔너리 리스트."""
    sheet = next((wb[s] for s in wb.sheetnames if name_fragment in s), None)
    if sheet is None:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    # 헤더 행 탐색 (첫 셀이 None이 아닌 행)
    header_idx = next((i for i, r in enumerate(rows) if r and r[0] is not None), 0)
    headers = [str(h).strip() if h is not None else f"col{j}"
               for j, h in enumerate(rows[header_idx])]
    result = []
    for row in rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        result.append(d)
    return result


# ---------------------------------------------------------------------------
# Price_Raw(인포) 직독 가격맵 + 이름 정규화
# ---------------------------------------------------------------------------

def _normalize_name(s: str) -> str:
    """공백/괄호/특수문자 제거한 매칭 키."""
    import re
    return re.sub(r"[\s()（）\[\]·,./-]", "", str(s)).lower()


def _to_date_str(v: Any) -> str | None:
    """Price_Raw 날짜 셀 → 'YYYY-MM-DD' 문자열."""
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]


def _latest_in_column(data_rows: list, c: int) -> tuple[Any, Any]:
    """컬럼 c를 아래에서 위로 역방향 탐색해 (날짜셀, 값) 최신 비결측값 반환."""
    for r in reversed(data_rows):
        v = r[c] if c < len(r) else None
        if v is not None:
            return (r[0] if r else None, v)
    return (None, None)


def _build_price_map() -> tuple[dict[str, dict], dict[str, Any]]:
    """Price_Raw(인포)에서 자산명→{price, kind, usd, price_date} 최신값 맵 생성.

    핵심 수정(P0-4): 전역 마지막 1행만 보지 않고, **컬럼별로 아래에서 위로
    역방향 탐색하여 각 자산의 최신 비결측값**을 가격으로 사용한다(트레일링 null 복구).
    price_kind도 동일 컬럼 기준. USD 표기 자산(자산마스터 infomax_type=FRN 또는
    USD/달러 표기)은 USDKRW 최신값으로 원화 환산한다.

    반환: (price_map, meta) — meta는 {"price_as_of", "usdkrw", "usdkrw_date"}.
    Uses _load_sheets_cached() — no workbook object kept open.
    """
    meta: dict[str, Any] = {"price_as_of": None, "usdkrw": None, "usdkrw_date": None}

    sheets = _load_sheets_cached()
    if not sheets:
        return {}, meta

    # USD 자산 식별: 자산마스터 infomax_type == FRN 의 report/internal 이름 집합
    usd_keys: set[str] = set()
    try:
        master_rows = next((v for k, v in sheets.items() if "자산마스터" in k), [])
        if master_rows:
            # parse header row
            hdr_idx = next((i for i, r in enumerate(master_rows) if r and r[0] is not None), 0)
            hdrs = [str(h).strip() if h is not None else f"col{j}"
                    for j, h in enumerate(master_rows[hdr_idx])]
            for mrow in master_rows[hdr_idx + 1:]:
                if not mrow or all(v is None for v in mrow):
                    continue
                rd = {hdrs[i]: mrow[i] for i in range(min(len(hdrs), len(mrow)))}
                itype = str(rd.get("infomax_type") or "").strip().upper()
                if itype == "FRN":
                    for nm in (rd.get("report_name"), rd.get("internal_name")):
                        if nm:
                            usd_keys.add(_normalize_name(nm))
    except Exception:
        pass

    rows = next((v for k, v in sheets.items() if "Price_Raw(인포)" in k), [])
    if len(rows) < 6:
        return {}, meta
    name_row = rows[3]
    type_row = rows[4]
    # 실제 데이터 행: 헤더(0~4) 이후 + 날짜(col0)가 채워진 행만
    data_rows = [r for r in rows[6:] if r and r[0] is not None]
    if not data_rows:
        return {}, meta

    # 시세 기준일(price_as_of): 데이터가 1개라도 존재하는 가장 마지막 날짜
    for r in reversed(data_rows):
        if any(v is not None for v in r[1:]):
            meta["price_as_of"] = _to_date_str(r[0])
            break

    ncol = len(name_row)

    # USDKRW 컬럼 최신값 탐색(있으면) — 원화 환산용
    usdkrw = None
    usdkrw_date = None
    for c in range(1, ncol):
        nm = name_row[c] if c < len(name_row) else None
        if nm and "USDKRW" in str(nm).upper().replace(" ", ""):
            dcell, val = _latest_in_column(data_rows, c)
            try:
                usdkrw = float(val) if val is not None else None
                usdkrw_date = _to_date_str(dcell)
            except (TypeError, ValueError):
                usdkrw = None
            break
    if usdkrw is None:
        # 대용: pykrx/네이버 'USDKRW=X' (best-effort, 무리면 None 유지)
        usdkrw, usdkrw_date = _fetch_usdkrw_fallback()
    meta["usdkrw"] = usdkrw
    meta["usdkrw_date"] = usdkrw_date

    price_map: dict[str, dict] = {}
    for c in range(1, ncol):
        raw_name = (name_row[c] if c < len(name_row) and name_row[c]
                    else (name_row[c-1] if c-1 < len(name_row) else None))
        if not raw_name:
            continue
        nkey = _normalize_name(raw_name)
        # USDKRW 컬럼 자체는 가격맵에서 제외
        if "USDKRW" in str(raw_name).upper().replace(" ", ""):
            continue
        dcell, val = _latest_in_column(data_rows, c)
        if val is None:
            continue
        try:
            price = float(val)
        except (TypeError, ValueError):
            continue
        kind = str(type_row[c]) if c < len(type_row) and type_row[c] else "현재가"
        # USD 자산 판별: 자산마스터 FRN 키 매칭 또는 이름에 USD/달러 표기
        nm_up = str(raw_name).upper()
        is_usd = (nkey in usd_keys) or ("USD" in nm_up) or ("달러" in str(raw_name))
        krw_price = price
        usd_converted = False
        if is_usd and usdkrw:
            krw_price = price * usdkrw
            usd_converted = True
        price_map[nkey] = {
            "name": str(raw_name),
            "price": krw_price,
            "price_native": price,
            "kind": kind,
            "usd": is_usd,
            "usd_converted": usd_converted,
            "price_date": _to_date_str(dcell),
        }
    return price_map, meta


def _fetch_usdkrw_fallback() -> tuple[float | None, str | None]:
    """USDKRW 환율 대용 조회(best-effort). 실패 시 (None, None)."""
    # 네이버 'USDKRW=X' (yfinance/네이버 금융) — 무리면 조용히 포기
    try:
        from .price_service import get_quote as _gq  # type: ignore
        q = _gq("USDKRW=X")
        if q and q.get("price"):
            return float(q["price"]), None
    except Exception:
        pass
    return None, None


def _master_krx_code(value: Any) -> str | None:
    """????? infomax_price_code ?? 6?? KRX ??? ??."""
    if value is None:
        return None
    text = str(value).strip().split('.')[0]
    if text.isdigit() and 1 <= len(text) <= 6:
        return text.zfill(6)
    return None


def _master_foreign_ticker(*values: Any) -> str | None:
    """자산마스터의 NAS:GRAB / GRAB US Equity 같은 값에서 해외 티커 추출."""
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        up = text.upper()
        if ":" in up:
            market, ticker = up.split(":", 1)
            ticker = ticker.strip().split()[0]
            if market in {"NAS", "NYS", "ASE", "AMEX", "NYSE", "NASDAQ"} and ticker:
                return ticker
        if up.endswith(" US EQUITY"):
            ticker = up.split()[0]
            if ticker:
                return ticker
    return None


def _build_asset_master(wb) -> dict[str, dict]:
    """internal_name/report_name ??? ? ? {report_name, bm_name, krx_code}."""
    master = _sheet_to_dicts(wb, "\uc790\uc0b0\ub9c8\uc2a4\ud130")
    out: dict[str, dict] = {}
    for row in master:
        report = row.get("report_name") or ""
        internal = row.get("internal_name") or ""
        bm = row.get("bm_name")
        krx_code = _master_krx_code(row.get("infomax_price_code"))
        foreign_ticker = _master_foreign_ticker(
            row.get("infomax_price_code"),
            row.get("bloomberg_price_ticker"),
            row.get("bloomberg_tr_ticker"),
        )
        entry = {
            "report_name": str(report),
            "bm_name": str(bm) if bm else None,
            "krx_code": krx_code,
            "foreign_ticker": foreign_ticker,
        }
        for key in (internal, report):
            if key:
                out[_normalize_name(key)] = entry
    return out

def _resolve_price(holding_name: str, price_map: dict[str, dict],
                   asset_master: dict[str, dict]) -> dict | None:
    """보유종목명 → Price_Raw 가격. 직접 매칭 → 자산마스터 경유 → 부분일치 순."""
    import re as _re
    key = _normalize_name(holding_name)
    if key in price_map:
        return price_map[key]
    am = asset_master.get(key)
    if am:
        rk = _normalize_name(am["report_name"])
        if rk in price_map:
            return price_map[rk]
        # AM report_name이 price_map 키의 prefix/substring인 경우도 시도
        if rk:
            for pk, pv in price_map.items():
                if rk in pk or pk in rk:
                    return pv
    # holding 키 부분일치 fallback
    for pk, pv in price_map.items():
        if key and (key in pk or pk in key):
            return pv
    # 장기무/호F 등 suffix 제거 후 재시도
    stripped = _re.sub(r"(장기무|장기|호f|호)$", "", key)
    if stripped and stripped != key:
        for pk, pv in price_map.items():
            if stripped in pk or pk in stripped:
                return pv
    # 앞 8자 이상 공통 prefix로 매칭 (펀드 약칭 처리)
    if len(key) >= 8:
        prefix = key[:8]
        for pk, pv in price_map.items():
            if pk.startswith(prefix):
                return pv
    # 공통 내부 토큰(6자 이상) 매칭 — TIME→TIMEFOLIO 등 브랜드 약칭 처리
    if len(key) >= 6:
        # key에서 앞 브랜드(영문+숫자) 제거 후 나머지가 price_map 키에 포함되면 매칭
        body = _re.sub(r"^[a-z0-9]+", "", key)
        if len(body) >= 6:
            for pk, pv in price_map.items():
                if body in pk:
                    return pv
    return None


# ---------------------------------------------------------------------------
# KRX 상장 ETF/주식 이름 -> 코드 매핑 (live_price 용, best-effort)
# ---------------------------------------------------------------------------

_krx_name_to_code: dict[str, str] | None = None


def _build_krx_name_to_code() -> dict[str, str]:
    """pykrx 의 ETF + 보통주 티커 목록으로 정규화 이름 -> 코드 맵 생성. 1회 캐시."""
    global _krx_name_to_code
    if _krx_name_to_code is not None:
        return _krx_name_to_code
    out: dict[str, str] = {}
    try:
        from pykrx import stock as krx
        today = dt.date.today().strftime("%Y%m%d")
        try:
            for code in krx.get_etf_ticker_list(today):
                try:
                    nm = krx.get_etf_ticker_name(code)
                    if nm:
                        out[_normalize_name(nm)] = code
                except Exception:
                    continue
        except Exception:
            pass
        for market in ("KOSPI", "KOSDAQ"):
            try:
                for code in krx.get_market_ticker_list(today, market=market):
                    try:
                        nm = krx.get_market_ticker_name(code)
                        if nm:
                            out.setdefault(_normalize_name(nm), code)
                    except Exception:
                        continue
            except Exception:
                pass
    except Exception:
        pass
    _krx_name_to_code = out
    return out


def _resolve_code(
    holding_name: str,
    name_to_code: dict[str, str],
    asset_master: dict[str, dict] | None = None,
) -> str | None:
    """????? -> KRX ??. ????? ?? ? pykrx ?? ?? ?."""
    key = _normalize_name(holding_name)
    if asset_master:
        am = asset_master.get(key)
        if am and am.get("krx_code"):
            return str(am["krx_code"])
    if not name_to_code:
        return None
    if key in name_to_code:
        return name_to_code[key]
    # ???? ????? ???(??? suffix) ?? ?? ????
    for nm_key, code in name_to_code.items():
        if len(nm_key) >= 6 and (nm_key in key or key in nm_key):
            return code
    return None


def _resolve_foreign_ticker(holding_name: str, asset_master: dict[str, dict] | None = None) -> str | None:
    """보유명 -> 해외 티커(best-effort)."""
    if not asset_master:
        return None
    key = _normalize_name(holding_name)
    am = asset_master.get(key)
    if am and am.get("foreign_ticker"):
        return str(am["foreign_ticker"])
    for nm_key, entry in asset_master.items():
        if key and (key in nm_key or nm_key in key) and entry.get("foreign_ticker"):
            return str(entry["foreign_ticker"])
    return None


def _get_yahoo_quote_usd(ticker: str) -> dict[str, Any] | None:
    """Yahoo chart API로 해외주식 USD 현재가 조회(best-effort)."""
    ticker = str(ticker or "").strip().upper()
    if not ticker:
        return None
    try:
        import requests
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        r = requests.get(
            url,
            params={"range": "5d", "interval": "1d"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=6,
        )
        r.raise_for_status()
        result = (r.json().get("chart", {}).get("result") or [None])[0]
        if not result:
            return None
        meta = result.get("meta") or {}
        price = meta.get("regularMarketPrice")
        prev = meta.get("previousClose")
        ts = meta.get("regularMarketTime")
        if price is None:
            closes = (((result.get("indicators") or {}).get("quote") or [{}])[0].get("close") or [])
            vals = [float(v) for v in closes if v is not None]
            if vals:
                price = vals[-1]
                if len(vals) >= 2:
                    prev = vals[-2]
        if price is None:
            return None
        price = float(price)
        change_pct = None
        if prev:
            prev = float(prev)
            if prev:
                change_pct = (price / prev - 1.0) * 100
        as_of = None
        if ts:
            as_of = dt.datetime.fromtimestamp(int(ts), tz=dt.timezone.utc).date().isoformat()
        return {
            "price": price,
            "change_pct": round(change_pct, 2) if change_pct is not None else None,
            "source": "YAHOO",
            "as_of": as_of,
            "ticker": ticker,
            "currency": "USD",
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# P1: 일별 시계열(가격/BM) 빌더 + BM 매핑 (P0 헬퍼 재사용)
# ---------------------------------------------------------------------------

DEFAULT_BM_NAME = "KOSPI"  # 미매핑 종목 디폴트 BM
TRADING_DAYS = 252         # 연율화 계수(영업일)


def _build_price_series() -> tuple[list[str], dict[str, dict]]:
    """Price_Raw(인포)에서 자산별 **일별 시계열**을 구축.

    P0의 _build_price_map 은 컬럼별 '최신 1값'만 뽑지만, 곡선/위험지표는 전 구간이
    필요하므로 동일한 컬럼/병합셀 규칙으로 일별 시계열을 만든다.

    반환:
      dates  : 'YYYY-MM-DD' 정렬 리스트(공통 일자축, col0 기준 오름차순)
      series : {정규화이름: {"name", "kind", "values": {date_str: float}}}
    Uses _load_sheets_cached() — no workbook object kept open.
    """
    rows = _get_sheet_rows("Price_Raw(인포)")
    if len(rows) < 7:
        return [], {}
    name_row = rows[3]
    type_row = rows[4]
    data_rows = [r for r in rows[6:] if r and r[0] is not None]
    if not data_rows:
        return [], {}

    dates: list[str] = []
    for r in data_rows:
        ds = _to_date_str(r[0])
        if ds:
            dates.append(ds)

    ncol = len(name_row)
    series: dict[str, dict] = {}
    for c in range(1, ncol):
        raw_name = (name_row[c] if c < len(name_row) and name_row[c]
                    else (name_row[c - 1] if c - 1 < len(name_row) else None))
        if not raw_name:
            continue
        if "USDKRW" in str(raw_name).upper().replace(" ", ""):
            continue
        nkey = _normalize_name(raw_name)
        kind = str(type_row[c]) if c < len(type_row) and type_row[c] else "현재가"
        vals: dict[str, float] = {}
        for r in data_rows:
            ds = _to_date_str(r[0])
            v = r[c] if c < len(r) else None
            if ds is None or v is None:
                continue
            try:
                vals[ds] = float(v)
            except (TypeError, ValueError):
                continue
        if vals:
            series[nkey] = {"name": str(raw_name), "kind": kind, "values": vals}
    return dates, series


_BM_ALIAS = {
    # 자산마스터 bm_name 정규화키 → BM_RAW 정규화키 보정(한글/영문 표기 차이)
    "코스닥150": "kosdaq150",
    "리츠top10": "krx리츠top10지수",
}


def _build_bm_series() -> dict[str, dict]:
    """BM_RAW(인포)에서 BM별 일별 시계열 맵.

    구조: row1=BM명(짝수 col), row2='일자'/'현재가', row3+=(일자,값) 컬럼쌍.
    반환: {정규화BM키: {"name", "values": {date_str: float}}}
    Uses _load_sheets_cached() — no workbook object kept open.
    """
    rows = _get_sheet_rows("BM_RAW(인포)")
    if len(rows) < 4:
        return {}
    name_row = rows[1]
    data_rows = rows[3:]
    out: dict[str, dict] = {}
    ncol = len(name_row)
    for c in range(0, ncol, 2):
        bm_name = name_row[c] if c < len(name_row) else None
        if not bm_name:
            continue
        nkey = _normalize_name(bm_name)
        vals: dict[str, float] = {}
        for r in data_rows:
            d = r[c] if c < len(r) else None
            v = r[c + 1] if c + 1 < len(r) else None
            if d is None or v is None:
                continue
            ds = _to_date_str(d)
            if ds is None:
                continue
            try:
                vals[ds] = float(v)
            except (TypeError, ValueError):
                continue
        if vals:
            out[nkey] = {"name": str(bm_name).strip(), "values": vals}
    return out


def _resolve_bm_series(bm_name: str | None, bm_series: dict[str, dict]) -> dict | None:
    """자산마스터 bm_name → BM_RAW 시계열. 직접 정규화 → 별칭 → 부분일치."""
    if not bm_name:
        return None
    key = _normalize_name(bm_name)
    if key in bm_series:
        return bm_series[key]
    aliased = _BM_ALIAS.get(key)
    if aliased and aliased in bm_series:
        return bm_series[aliased]
    for bk, bv in bm_series.items():
        if key and (key in bk or bk in key):
            return bv
    return None


def _bm_return_pct(values: dict[str, float], start_date: str | None) -> float | None:
    """BM 시계열 values에서 start_date 이후 첫 값 대비 최신 값 수익률(%)."""
    if not values:
        return None
    dates = sorted(values)
    start_key = _to_date_str(start_date) if start_date else None
    if start_key:
        base_key = next((d for d in dates if d >= start_key and values.get(d)), None)
    else:
        base_key = next((d for d in dates if values.get(d)), None)
    last_key = next((d for d in reversed(dates) if values.get(d)), None)
    if not base_key or not last_key:
        return None
    base = values.get(base_key)
    last = values.get(last_key)
    if not base or not last:
        return None
    return round((last / base - 1.0) * 100, 2)


def _first_value_on_or_after(values: dict[str, float], start_date: str | None) -> float | None:
    """시계열에서 start_date 이후 첫 유효값. 없으면 전체 첫 유효값."""
    if not values:
        return None
    dates = sorted(values)
    start_key = _to_date_str(start_date) if start_date else None
    if start_key:
        val = next((values[d] for d in dates if d >= start_key and values.get(d)), None)
        if val:
            return val
    return next((values[d] for d in dates if values.get(d)), None)


def _previous_value_before(values: dict[str, float], date_str: str | None) -> float | None:
    """시계열에서 date_str 직전의 유효값."""
    if not values or not date_str:
        return None
    key = _to_date_str(date_str)
    if not key:
        return None
    dates = sorted(values)
    return next((values[d] for d in reversed(dates) if d < key and values.get(d)), None)


def _latest_and_previous_values(
    values: dict[str, float],
    preferred_date: str | None = None,
) -> tuple[str | None, float | None, float | None]:
    """Return (latest_date, latest_value, previous_value) for daily return.

    값이 전일과 같다는 이유만으로 최신 행을 건너뛰면, 실제 보합일(0.00%)도
    전전일 대비 수익률로 잘못 계산된다. 날짜 기준으로 최신/직전 유효값을 고른다.
    """
    if not values:
        return None, None, None
    dates = sorted(d for d, v in values.items() if v is not None)
    if not dates:
        return None, None, None

    key = _to_date_str(preferred_date)
    if key:
        eligible = [d for d in dates if d <= key]
        latest_date = eligible[-1] if eligible else dates[-1]
    else:
        latest_date = dates[-1]

    idx = dates.index(latest_date)
    latest_value = values.get(latest_date)
    previous_value = values.get(dates[idx - 1]) if idx > 0 else None
    return latest_date, latest_value, previous_value


def _ffill(dates: list[str], values: dict[str, float]) -> list[float | None]:
    """공통 일자축 dates 에 대해 values 를 직전값 ffill 한 리스트 반환."""
    out: list[float | None] = []
    last: float | None = None
    for d in dates:
        if d in values:
            last = values[d]
        out.append(last)
    return out


def _daily_returns(series: list[float]) -> list[float]:
    """단순 일별수익률 리스트(길이 len-1). 0/None 가격은 건너뜀 처리."""
    rets: list[float] = []
    for i in range(1, len(series)):
        prev = series[i - 1]
        cur = series[i]
        if prev and cur and prev != 0:
            rets.append(cur / prev - 1.0)
        else:
            rets.append(0.0)
    return rets


def _mean(xs: list[float]) -> float:
    return sum(xs) / len(xs) if xs else 0.0


def _stdev(xs: list[float]) -> float:
    """표본표준편차(ddof=1)."""
    n = len(xs)
    if n < 2:
        return 0.0
    m = _mean(xs)
    var = sum((x - m) ** 2 for x in xs) / (n - 1)
    return math.sqrt(var)


def _max_drawdown(values: list[float]) -> float:
    """최대낙폭(MDD, 음수). values=누적 평가액/지수 시계열."""
    peak = None
    mdd = 0.0
    for v in values:
        if v is None or v <= 0:
            continue
        if peak is None or v > peak:
            peak = v
        if peak:
            dd = v / peak - 1.0
            if dd < mdd:
                mdd = dd
    return mdd


def _build_curve_context(period: str | None = None) -> dict:
    """곡선/위험지표 공통 컨텍스트(일자축·포트평가액·BM지수·커버리지) 1회 산출.

    반환 dict 키:
      dates, port_value(list), port_cost(float), bm_name, bm_index(list),
      bm_resolved(bool), coverage_pct, matched_value, total_value,
      bm_weight(dict), default_used(bool)
    """
    wb = _load_wb()
    holdings_raw = _sheet_to_dicts(wb, "현재보유") if wb else []
    asset_master = _build_asset_master(wb) if wb else {}
    # bm_name 매핑(report/internal 정규화키 → bm_name) — 자산마스터 직접 파싱
    bm_by_key: dict[str, str] = {}
    if wb is not None:
        for row in _sheet_to_dicts(wb, "자산마스터"):
            bm = row.get("bm_name")
            for nm in (row.get("internal_name"), row.get("report_name")):
                if nm and bm:
                    bm_by_key[_normalize_name(nm)] = str(bm)
        wb.close()

    dates, price_series = _build_price_series()
    bm_series = _build_bm_series()
    price_map, _meta = _build_price_map()

    def col(d: dict, *cands: str) -> str | None:
        for k in d:
            kk = str(k).replace(" ", "")
            if any(c.replace(" ", "") in kk for c in cands):
                return k
        return None

    h0 = holdings_raw[0] if holdings_raw else {}
    name_c = col(h0, "종목", "종 목")
    qty_c = col(h0, "주수", "수량")
    cost_c = col(h0, "투자원금", "원금")

    # period → 최근 N영업일 슬라이스 인덱스
    if period:
        p = str(period).upper()
        n_map = {"1M": 21, "3M": 63, "6M": 126, "1Y": 252, "YTD": None}
        if p in ("1Y", "1년"):
            keep = 252
        elif p in ("3M", "3개월"):
            keep = 63
        elif p in ("6M", "6개월"):
            keep = 126
        elif p in ("1M", "1개월"):
            keep = 21
        elif p in ("MAX", "ALL", "전체"):
            keep = None
        else:
            keep = None
        if keep and len(dates) > keep:
            dates = dates[-keep:]

    n = len(dates)
    port_value = [0.0] * n
    port_cost = 0.0
    matched_value = 0.0
    total_value = 0.0
    bm_weight: dict[str, float] = {}    # bm_name → 평가액 가중
    coverage_matched = 0.0              # bm 매핑된 평가액
    default_used = False

    for h in holdings_raw:
        name = str(h.get(name_c, "")).strip() if name_c else ""
        if not name:
            continue
        qty = float(h.get(qty_c, 0) or 0) if qty_c else 0.0
        cost = float(h.get(cost_c, 0) or 0) if cost_c else 0.0
        port_cost += cost

        # 가격 시계열 매칭: 직접 → 자산마스터 report_name 경유 → 부분일치
        pinfo = _resolve_price(name, price_map, asset_master)
        ts = None
        if pinfo:
            pkey = _normalize_name(pinfo["name"])
            ts = price_series.get(pkey)
        if ts is None:
            # 직접 시계열 키 탐색(부분일치)
            nk = _normalize_name(name)
            for sk, sv in price_series.items():
                if nk and (nk in sk or sk in nk):
                    ts = sv
                    break
        latest_val = 0.0
        if ts is not None:
            filled = _ffill(dates, ts["values"])
            for i, v in enumerate(filled):
                if v is not None:
                    port_value[i] += v * qty
            last = next((v for v in reversed(filled) if v is not None), None)
            if last is not None:
                latest_val = last * qty
        total_value += latest_val

        # BM 매핑(평가액 가중 다수결용)
        nk = _normalize_name(name)
        bm_name = bm_by_key.get(nk)
        if bm_name is None:
            am = asset_master.get(nk)
            if am and am.get("bm_name"):
                bm_name = am["bm_name"]
        if bm_name is None:
            for bk, bv in bm_by_key.items():
                if nk and (nk in bk or bk in nk):
                    bm_name = bv
                    break
        weight = latest_val if latest_val > 0 else cost
        if bm_name:
            coverage_matched += weight
            bm_weight[bm_name] = bm_weight.get(bm_name, 0.0) + weight
            matched_value += weight
        else:
            default_used = True
            bm_weight[DEFAULT_BM_NAME] = bm_weight.get(DEFAULT_BM_NAME, 0.0) + weight

    denom = sum(bm_weight.values()) or 1.0
    coverage_pct = round(coverage_matched / denom * 100, 1)

    # 손익현황의 VS 위험지표/손익 시계열은 자산군 다수결 BM이 아니라
    # 시장 대표지수인 KOSPI를 고정 비교 기준으로 사용한다.
    # 보유종목별 BM은 테이블 노출용으로 유지하되, 포트폴리오 위험/곡선 BM에는 쓰지 않는다.
    rep_bm = DEFAULT_BM_NAME
    bm_ts = _resolve_bm_series(rep_bm, bm_series)
    bm_resolved = bm_ts is not None
    if bm_ts is None:
        bm_ts = _resolve_bm_series(DEFAULT_BM_NAME, bm_series)
        if bm_ts is not None:
            rep_bm = DEFAULT_BM_NAME
            default_used = True

    # 포트 시작=100 리베이스, BM도 동일 시작점 100 리베이스
    bm_filled = _ffill(dates, bm_ts["values"]) if bm_ts else [None] * n
    base_port = next((v for v in port_value if v and v > 0), None)
    base_bm = next((v for v in bm_filled if v and v > 0), None)

    return {
        "dates": dates,
        "port_value": port_value,
        "port_cost": port_cost,
        "bm_name": (bm_ts["name"] if bm_ts else rep_bm),
        "bm_filled": bm_filled,
        "base_port": base_port,
        "base_bm": base_bm,
        "bm_resolved": bm_resolved,
        "coverage_pct": coverage_pct,
        "matched_value": matched_value,
        "total_value": total_value,
        "bm_weight": bm_weight,
        "default_used": default_used,
    }


def get_pnl_curve(period: str | None = None) -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_curve
        return mock_pnl_curve(period)
    """손익 시계열 곡선 + BM 오버레이.

    반환:
      dates: ['YYYY-MM-DD', ...]
      portfolio_value: [평가액(원), ...]  (보유 주수 × 일별가격 합산, 결측 ffill)
      portfolio_index: [100 기준 리베이스, ...]
      cum_pnl: [누적손익=평가-원금, ...]
      portfolio_cost: 투자원금 합계(number)
      bm_index: [BM 100 리베이스(포트 시작점과 동일), ...]
      bm_name, as_of, days, coverage_pct, bm_resolved
    """
    ctx = _build_curve_context(period)
    dates = ctx["dates"]
    port_value = ctx["port_value"]
    port_cost = ctx["port_cost"]
    base_port = ctx["base_port"]
    base_bm = ctx["base_bm"]
    bm_filled = ctx["bm_filled"]

    portfolio_index: list[float | None] = []
    for v in port_value:
        if base_port and v:
            portfolio_index.append(round(v / base_port * 100.0, 4))
        else:
            portfolio_index.append(None)

    bm_index: list[float | None] = []
    for v in bm_filled:
        if base_bm and v:
            bm_index.append(round(v / base_bm * 100.0, 4))
        else:
            bm_index.append(None)

    cum_pnl = [round(v - port_cost, 0) if v else None for v in port_value]

    # realized_pnl_total: 매도내역 PR 합산 (억원 단위 → 원)
    wb2 = _load_wb()
    sells_raw2 = _sheet_to_dicts(wb2, "매도내역") if wb2 else []
    if wb2:
        wb2.close()
    realized_total = _calc_realized_pnl_total(sells_raw2)
    # realized_cum: 날짜별 실현 누적 — 매도내역 날짜 파싱 후 날짜 기준 누적
    # 거친 데이터(날짜 불명확)이면 상수(realized_total)로 모든 날 동일 값으로 채움
    realized_cum = _build_realized_cum(dates, sells_raw2, realized_total)
    # total_incl_realized: 미실현(cum_pnl) + 실현손익
    total_incl_realized = [
        round(c + realized_total, 0) if c is not None else None
        for c in cum_pnl
    ]

    return {
        "dates": dates,
        "portfolio_value": [round(v, 0) if v else 0.0 for v in port_value],
        "portfolio_index": portfolio_index,
        "cum_pnl": cum_pnl,
        "portfolio_cost": round(port_cost, 0),
        "bm_index": bm_index,
        "bm_name": ctx["bm_name"],
        "bm_resolved": ctx["bm_resolved"],
        "coverage_pct": ctx["coverage_pct"],
        "realized_cum": realized_cum,
        "total_incl_realized": total_incl_realized,
        "realized_pnl_total": realized_total,
        "days": len(dates),
        "period": period or "MAX",
        "as_of": dates[-1] if dates else None,
        "fetched_at": dt.datetime.now().isoformat(timespec="seconds"),
    }


def get_pnl_risk(period: str | None = None) -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_risk
        return mock_pnl_risk(period)
    """위험지표: 연율수익/변동성/MDD/베타/TE/IR/초과수익 + 정합성 메타.

    산식(methodology 에 명시):
      - 일별수익률 r_p(포트), r_b(BM) = P_t/P_(t-1) - 1
      - 연율수익 = mean(r) * 252
      - 연율변동성 = stdev(r, ddof=1) * sqrt(252)
      - MDD = min(누적/직전고점 - 1)  (음수)
      - beta = cov(r_p, r_b) / var(r_b)   (전 구간 회귀)
      - TE = stdev(r_p - r_b, ddof=1) * sqrt(252)
      - IR = (연율 r_p - 연율 r_b) / TE
      - 초과수익 = 연율 r_p - 연율 r_b
    """
    ctx = _build_curve_context(period)
    dates = ctx["dates"]
    port_value = ctx["port_value"]
    bm_filled = ctx["bm_filled"]

    # 포트/BM 모두 값이 있는 구간만 사용(정합성)
    p_clean: list[float] = []
    b_clean: list[float] = []
    for pv, bv in zip(port_value, bm_filled):
        if pv and pv > 0 and bv and bv > 0:
            p_clean.append(pv)
            b_clean.append(bv)

    rp = _daily_returns(p_clean)
    rb = _daily_returns(b_clean)
    nobs = min(len(rp), len(rb))
    rp = rp[:nobs]
    rb = rb[:nobs]

    ann_return = _mean(rp) * TRADING_DAYS
    ann_vol = _stdev(rp) * math.sqrt(TRADING_DAYS)
    bm_ann_return = _mean(rb) * TRADING_DAYS
    mdd = _max_drawdown(p_clean)

    # beta = cov(rp, rb) / var(rb)
    beta = 0.0
    if nobs >= 2:
        mp, mb = _mean(rp), _mean(rb)
        cov = sum((rp[i] - mp) * (rb[i] - mb) for i in range(nobs)) / (nobs - 1)
        varb = sum((rb[i] - mb) ** 2 for i in range(nobs)) / (nobs - 1)
        beta = cov / varb if varb else 0.0

    excess_daily = [rp[i] - rb[i] for i in range(nobs)]
    tracking_error = _stdev(excess_daily) * math.sqrt(TRADING_DAYS)
    excess_return = ann_return - bm_ann_return
    info_ratio = (excess_return / tracking_error) if tracking_error else 0.0

    # 정합성 sample_check: 단일종목 베타를 별도 손계산하여 메타 베타와 대조
    sample_check = _sample_beta_check(dates, ctx)

    methodology = (
        f"daily simple returns P_t/P_(t-1)-1; annualization factor={TRADING_DAYS} "
        f"(trading days); ann_return=mean(r)*{TRADING_DAYS}; "
        f"ann_vol=stdev(r,ddof=1)*sqrt({TRADING_DAYS}); MDD=min(V_t/peak-1); "
        f"beta=cov(r_p,r_b)/var(r_b) over full overlap window (n={nobs}); "
        f"TE=stdev(r_p-r_b,ddof=1)*sqrt({TRADING_DAYS}); "
        f"IR=(ann_r_p-ann_r_b)/TE; excess_return=ann_r_p-ann_r_b. "
        f"BM rebased to portfolio start=100."
    )

    # Sharpe = ann_return / ann_vol (risk-free = 0; both in decimal units)
    sharpe: float | None = None
    if ann_vol and abs(ann_vol) > 1e-9:
        sharpe = round(ann_return / ann_vol, 4)

    # Calmar = ann_return / |MDD| (MDD is negative decimal)
    calmar: float | None = None
    if mdd and abs(mdd) > 1e-9:
        calmar = round(ann_return / abs(mdd), 4)

    return {
        "ann_return": round(ann_return, 4),
        "ann_vol": round(ann_vol, 4),
        "mdd": round(mdd, 4),
        "beta": round(beta, 4),
        "tracking_error": round(tracking_error, 4),
        "info_ratio": round(info_ratio, 4),
        "excess_return": round(excess_return, 4),
        "bm_ann_return": round(bm_ann_return, 4),
        "sharpe": sharpe,
        "calmar": calmar,
        "bm_name": ctx["bm_name"],
        "bm_resolved": ctx["bm_resolved"],
        "coverage_pct": ctx["coverage_pct"],
        "default_bm_used": ctx["default_used"],
        "n_obs": nobs,
        "period": period or "MAX",
        "methodology": methodology,
        "rf_note": "risk_free=0 (sharpe/calmar); units=decimals (ann_return/ann_vol/mdd)",
        "sample_check": sample_check,
        "as_of": dates[-1] if dates else None,
        "fetched_at": dt.datetime.now().isoformat(timespec="seconds"),
    }


def _sample_beta_check(dates: list[str], ctx: dict) -> dict:
    """단일 보유종목 1개의 베타를 독립 경로로 손계산하여 API 베타 산식과 대조.

    포트 전체 베타가 아니라 '대표 BM 대비 단일종목 베타'를 별도로 계산해, 동일 산식이
    재현되는지(회귀 일관성)와 오차를 기록한다. 검증 실패 시 note 에 사유.
    """
    out: dict[str, Any] = {"asset": None, "beta_hand": None, "beta_formula": None,
                           "abs_error": None, "note": ""}
    try:
        _dates2, price_series = _build_price_series()
        if ctx.get("dates"):
            dates = ctx["dates"]
        bm_filled = ctx["bm_filled"]
        # 충분한 시계열을 가진 첫 종목 선택
        cand = None
        for sk, sv in price_series.items():
            filled = _ffill(dates, sv["values"])
            cnt = sum(1 for v in filled if v and v > 0)
            if cnt > max(30, len(dates) // 2):
                cand = (sk, sv, filled)
                break
        if cand is None:
            out["note"] = "no asset with sufficient series"
            return out
        sk, sv, filled = cand
        p_clean, b_clean = [], []
        for pv, bv in zip(filled, bm_filled):
            if pv and pv > 0 and bv and bv > 0:
                p_clean.append(pv)
                b_clean.append(bv)
        rp = _daily_returns(p_clean)
        rb = _daily_returns(b_clean)
        n = min(len(rp), len(rb))
        rp, rb = rp[:n], rb[:n]
        if n < 2:
            out["note"] = "insufficient overlap"
            return out
        # 손계산 베타(명시적 합 공식)
        sp = sum(rp) / n
        sb = sum(rb) / n
        cov = sum((rp[i] - sp) * (rb[i] - sb) for i in range(n)) / (n - 1)
        varb = sum((rb[i] - sb) ** 2 for i in range(n)) / (n - 1)
        beta_hand = cov / varb if varb else 0.0
        # 산식 재현(동일하지만 독립 호출 경로) — _stdev/상관 기반
        mp, mb = _mean(rp), _mean(rb)
        cov2 = sum((rp[i] - mp) * (rb[i] - mb) for i in range(n)) / (n - 1)
        var2 = _stdev(rb) ** 2
        beta_formula = cov2 / var2 if var2 else 0.0
        out.update({
            "asset": sv["name"],
            "n_obs": n,
            "beta_hand": round(beta_hand, 6),
            "beta_formula": round(beta_formula, 6),
            "abs_error": round(abs(beta_hand - beta_formula), 8),
            "note": "single-asset beta vs BM; hand-sum vs formula path",
        })
    except Exception as e:  # pragma: no cover
        out["note"] = f"sample_check error: {e}"
    return out


# ---------------------------------------------------------------------------
# Attribution + Rolling Risk
# ---------------------------------------------------------------------------

def get_pnl_attribution() -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_attribution
        return mock_pnl_attribution()
    """자산군별 포트폴리오 기여 분석.

    bm_name(자산마스터) 기준으로 현재 보유를 그룹화하여 그룹별 시장가치·손익·기여도를 반환.
    반환: {groups:[{group, weight_pct, market_value, pnl, pnl_contribution_pct,
                    avg_return_pct, holdings_count}], total_market_value, total_pnl, as_of}
    """
    try:
        wb = _load_wb()
        if wb is None:
            return {"groups": [], "error": "Excel 파일 없음", "total_market_value": 0, "total_pnl": 0}

        holdings_raw = _sheet_to_dicts(wb, "현재보유")
        asset_master = _build_asset_master(wb)

        # bm_name 매핑: internal_name/report_name 정규화키 → bm_name
        bm_by_key: dict[str, str] = {}
        for row in _sheet_to_dicts(wb, "자산마스터"):
            bm = row.get("bm_name")
            for nm in (row.get("internal_name"), row.get("report_name")):
                if nm and bm:
                    bm_by_key[_normalize_name(nm)] = str(bm)
        wb.close()

        price_map, _meta = _build_price_map()

        def col(d: dict, *cands: str) -> str | None:
            for k in d:
                kk = str(k).replace(" ", "")
                if any(c.replace(" ", "") in kk for c in cands):
                    return k
            return None

        if not holdings_raw:
            return {"groups": [], "error": "현재보유 시트 비어있음", "total_market_value": 0, "total_pnl": 0}

        h0 = holdings_raw[0]
        name_c = col(h0, "종목", "종 목")
        qty_c  = col(h0, "주수", "수량")
        cost_c = col(h0, "투자원금", "원금")
        unit_c = col(h0, "취득단가", "단가")

        # 그룹별 집계: group → {market_value, cost, count}
        groups_agg: dict[str, dict] = {}

        for h in holdings_raw:
            name = str(h.get(name_c, "")).strip() if name_c else ""
            if not name:
                continue
            qty  = float(h.get(qty_c, 0) or 0) if qty_c else 0.0
            cost = float(h.get(cost_c, 0) or 0) if cost_c else 0.0
            unit = float(h.get(unit_c, 0) or 0) if unit_c else 0.0

            pinfo = _resolve_price(name, price_map, asset_master)
            cur_price = pinfo["price"] if pinfo else 0.0
            market_value = cur_price * qty if (cur_price and qty) else 0.0

            # cost 복원 (해외주식 fallback)
            if not cost and unit and qty:
                is_usd = bool(pinfo.get("usd_converted")) if pinfo else False
                usdkrw = _meta.get("usdkrw") or 1.0
                cost = unit * qty * (usdkrw if is_usd else 1.0)

            pnl = market_value - cost if cost else 0.0

            # BM 그룹 결정
            nk = _normalize_name(name)
            group = bm_by_key.get(nk)
            if group is None:
                am = asset_master.get(nk)
                if am and am.get("bm_name"):
                    group = am["bm_name"]
            if group is None:
                for bk, bv in bm_by_key.items():
                    if nk and (nk in bk or bk in nk):
                        group = bv
                        break
            if group is None:
                group = "기타/미분류"

            if group not in groups_agg:
                groups_agg[group] = {"market_value": 0.0, "cost": 0.0, "count": 0, "pnl_items": []}
            groups_agg[group]["market_value"] += market_value
            groups_agg[group]["cost"] += cost
            groups_agg[group]["count"] += 1
            if cost:
                groups_agg[group]["pnl_items"].append(pnl / cost * 100)

        total_market_value = sum(g["market_value"] for g in groups_agg.values()) or 1.0
        total_pnl = sum(g["market_value"] - g["cost"] for g in groups_agg.values())

        # pnl_contribution_pct 분모: total_pnl with abs fallback (same as summary)
        pnl_denom = total_pnl if abs(total_pnl) > 1e-9 else (
            sum(abs(g["market_value"] - g["cost"]) for g in groups_agg.values()) or 1.0
        )

        groups_out = []
        for group, g in groups_agg.items():
            mv = g["market_value"]
            pnl = mv - g["cost"]
            avg_ret = sum(g["pnl_items"]) / len(g["pnl_items"]) if g["pnl_items"] else 0.0
            groups_out.append({
                "group": group,
                "weight_pct": round(mv / total_market_value * 100, 2),
                "market_value": round(mv, 0),
                "pnl": round(pnl, 0),
                "pnl_contribution_pct": round(pnl / pnl_denom * 100, 2),
                "avg_return_pct": round(avg_ret, 2),
                "holdings_count": g["count"],
            })

        groups_out.sort(key=lambda x: x["market_value"], reverse=True)

        return {
            "groups": groups_out,
            "total_market_value": round(total_market_value, 0),
            "total_pnl": round(total_pnl, 0),
            "as_of": dt.datetime.now().isoformat(timespec="seconds"),
        }
    except Exception as e:
        return {"groups": [], "error": str(e), "total_market_value": 0, "total_pnl": 0}


def get_pnl_rolling_risk(window: int = 60) -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_rolling_risk
        return mock_pnl_rolling_risk(window)
    """롤링 위험지표(베타·IR) 시계열.

    _build_curve_context 와 동일한 포트/BM 시계열을 재사용하여
    window-일 롤링 베타(cov/var)와 롤링 IR(annualized excess/std)을 계산.

    반환: {dates, beta, ir, window, as_of}
    """
    try:
        ctx = _build_curve_context(None)  # 전체 기간
        port_value = ctx["port_value"]
        bm_filled  = ctx["bm_filled"]
        dates      = ctx["dates"]

        # 포트/BM 모두 유효한 구간 정렬
        p_clean: list[float] = []
        b_clean: list[float] = []
        d_clean: list[str]   = []
        for d, pv, bv in zip(dates, port_value, bm_filled):
            if pv and pv > 0 and bv and bv > 0:
                p_clean.append(pv)
                b_clean.append(bv)
                d_clean.append(d)

        rp = _daily_returns(p_clean)
        rb = _daily_returns(b_clean)
        n = min(len(rp), len(rb))
        rp = rp[:n]
        rb = rb[:n]
        # 수익률 날짜: d_clean[1:] (수익률은 t−1→t 이므로 길이 = n)
        ret_dates = d_clean[1: n + 1]

        out_dates: list[str] = []
        out_beta:  list[float | None] = []
        out_ir:    list[float | None] = []

        for i in range(n):
            start = max(0, i - window + 1)
            wp = rp[start: i + 1]
            wb_w = rb[start: i + 1]
            wlen = len(wp)
            # 날짜 라벨
            out_dates.append(ret_dates[i])

            if wlen < 4:
                out_beta.append(None)
                out_ir.append(None)
                continue

            # 롤링 베타
            mp, mb = _mean(wp), _mean(wb_w)
            cov  = sum((wp[j] - mp) * (wb_w[j] - mb) for j in range(wlen)) / (wlen - 1)
            varb = sum((wb_w[j] - mb) ** 2 for j in range(wlen)) / (wlen - 1)
            beta_r = round(cov / varb, 4) if varb else None

            # 롤링 IR: annualized (mean(excess)/std(excess)) × √252
            excess = [wp[j] - wb_w[j] for j in range(wlen)]
            ex_std = _stdev(excess)
            ex_mean = _mean(excess)
            if ex_std and abs(ex_std) > 1e-12:
                ir_r = round((ex_mean / ex_std) * math.sqrt(TRADING_DAYS), 4)
            else:
                ir_r = None

            out_beta.append(beta_r)
            out_ir.append(ir_r)

        return {
            "dates": out_dates,
            "beta": out_beta,
            "ir": out_ir,
            "window": window,
            "as_of": d_clean[-1] if d_clean else None,
        }
    except Exception as e:
        return {"dates": [], "beta": [], "ir": [], "window": window, "error": str(e)}


# ---------------------------------------------------------------------------
# Realized P&L helpers  (Task B: header-sniffed column map)
# ---------------------------------------------------------------------------

def _col_sniff(d: dict, *cands: str) -> str | None:
    """헤더-값 딕셔너리에서 후보 키워드 중 하나를 포함하는 첫 번째 키 반환.

    공백 제거 후 비교. 현재보유 시트에서 쓰는 col() 와 동일 로직.
    """
    for k in d:
        kk = str(k).replace(" ", "")
        if any(c.replace(" ", "") in kk for c in cands):
            return k
    return None


def _sniff_sells_cols(sells_raw: list[dict]) -> tuple[dict[str, str | None], bool]:
    """매도내역 헤더 행에서 컬럼명을 스니핑하여 논리키→실제컬럼 맵 반환.

    반환: (col_map, fallback_used)
      col_map keys: "name", "date", "qty", "buy_price", "sell_price", "pr", "tr"
      fallback_used: 헤더 스니핑이 부분적으로 실패해 위치 기반 폴백을 쓴 경우 True.

    위치 기반 폴백(vals[4]=PR, vals[5]=TR, vals[1]=date, vals[0]=name, vals[2]=qty):
    만약 헤더 스니핑 실패 시 None으로 마킹하고 fallback_used=True.
    """
    fallback_used = False
    col_map: dict[str, str | None] = {
        "name": None, "date": None, "qty": None,
        "buy_price": None, "sell_price": None,
        "pr": None, "tr": None,
    }
    if not sells_raw:
        return col_map, fallback_used

    h = sells_raw[0]  # 첫 행(실제 데이터 행)의 키가 헤더
    col_map["name"]       = _col_sniff(h, "종목", "명")
    col_map["date"]       = _col_sniff(h, "날짜", "매도일", "일자")
    col_map["qty"]        = _col_sniff(h, "수량")
    col_map["buy_price"]  = _col_sniff(h, "매입", "매수", "평단")
    col_map["sell_price"] = _col_sniff(h, "매도가", "체결")
    col_map["pr"]         = _col_sniff(h, "PR", "실현손익")
    col_map["tr"]         = _col_sniff(h, "TR", "수익률", "률")

    # 위치 기반 폴백: 스니핑 실패한 핵심 컬럼에만 적용
    # (sells_raw rows are dict keyed by header strings from _sheet_to_dicts)
    keys = list(h.keys())
    def _fallback(key: str, idx: int):
        nonlocal fallback_used
        if col_map[key] is None:
            if idx < len(keys):
                col_map[key] = keys[idx]
                fallback_used = True

    _fallback("name",  0)
    _fallback("date",  1)
    _fallback("qty",   2)
    _fallback("pr",    4)
    _fallback("tr",    5)

    return col_map, fallback_used


def _pr_value_krw(raw) -> float | None:
    """매도내역 PR 셀 값 → 원(₩) 단위 float. 억원 단위면 ×1억 환산.

    |PR| < 1000 이면 억원 단위로 간주(기존 heuristic 유지). 숫자 변환 실패 시 None.
    """
    if raw is None:
        return None
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return None
    if v == 0:
        return None
    if abs(v) < 1000:
        v = v * 1e8
    return v


def _tr_value_pct(raw) -> float | None:
    """매도내역 TR 셀 값 → 수익률(%). 숫자 변환 실패 시 None."""
    if raw is None:
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _build_realized_cum(dates: list[str], sells_raw: list[dict], fallback_total: float) -> list[float | None]:
    """날짜별 실현손익 누적 리스트.

    매도내역에서 날짜를 파싱해 날짜 기준 누적합을 구성한다.
    날짜 파싱 실패 종목이 과반이면 fallback_total 상수로 전체 채움.
    Task B: uses header-sniffed col_map instead of fixed positional indices.
    """
    if not dates:
        return []

    col_map, _fb = _sniff_sells_cols(sells_raw)

    daily_real: dict[str, float] = {}
    for t in sells_raw:
        raw_name = t.get(col_map["name"]) if col_map["name"] else None
        if not raw_name:
            continue
        raw_pr = t.get(col_map["pr"]) if col_map["pr"] else None
        v = _pr_value_krw(raw_pr)
        if v is None:
            continue
        raw_date = t.get(col_map["date"]) if col_map["date"] else None
        ds = _to_date_str(raw_date) if raw_date else None
        if ds:
            daily_real[ds] = daily_real.get(ds, 0.0) + v

    if not daily_real:
        return [fallback_total] * len(dates)
    cum = 0.0
    result: list[float | None] = []
    for d in dates:
        if d in daily_real:
            cum += daily_real[d]
        result.append(round(cum, 0))
    return result


def _calc_realized_pnl_total(sells_raw: list[dict]) -> float:
    """매도내역에서 PR 실현손익을 합산해 원(₩) 단위로 반환.

    Task B: uses header-sniffed col_map. Falls back to positional if headers absent.
    PR 값이 억원 단위이면 ×1억 환산. 판별: |PR| < 1000 이면 억원 단위로 간주.
    """
    col_map, _fb = _sniff_sells_cols(sells_raw)
    total = 0.0
    for t in sells_raw:
        raw_name = t.get(col_map["name"]) if col_map["name"] else None
        if not raw_name:
            continue
        raw_pr = t.get(col_map["pr"]) if col_map["pr"] else None
        v = _pr_value_krw(raw_pr)
        if v is not None:
            total += v
    return round(total, 0)


# ---------------------------------------------------------------------------
# Holding series (드릴다운 차트용)
# ---------------------------------------------------------------------------

def get_holding_series(key: str, period: str | None = None) -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_holding_series
        return mock_holding_series(key, period)
    """종목별 가격 시계열 + BM 시계열 (시작=100 리베이스).

    Parameters
    ----------
    key    : 종목명 또는 KRX 코드 (정규화 매칭)
    period : MAX(기본) | 1Y | 3M | 1M  (최근 N영업일 제한)

    Returns
    -------
    {name, dates, price_index, bm_index, bm_name, period, as_of}
    """
    _, price_series = _build_price_series()
    bm_all_series = _build_bm_series()

    # 자산마스터에서 bm_name 조회
    wb = _load_wb()
    asset_master = _build_asset_master(wb) if wb else {}
    if wb:
        wb.close()

    # key로 price_series 매칭 (정규화 직접 → 자산마스터 report_name 경유 → 부분일치)
    # _resolve_price 와 동일하게: 보유종목명(internal_name)이 Price_Raw 컬럼명과 다른
    # 펀드/리츠도 자산마스터의 report_name 으로 해석되도록 한다.
    nkey = _normalize_name(key)
    matched_ts: dict | None = None
    matched_name: str = key
    # 1) 정규화 직접
    if nkey in price_series:
        matched_ts = price_series[nkey]
    # 2) get_pnl_summary 와 동일한 5단계 _resolve_price 로 Price_Raw 컬럼 해석
    #    (펀드처럼 보유명/report_name/컬럼명이 모두 다른 변형명까지 매칭)
    if matched_ts is None:
        try:
            _pm, _ = _build_price_map()
            resolved = _resolve_price(key, _pm, asset_master)
            if resolved and resolved.get("name"):
                rk = _normalize_name(resolved["name"])
                if rk in price_series:
                    matched_ts = price_series[rk]
        except Exception:
            pass
    # 3) 자산마스터 report_name 경유
    if matched_ts is None:
        am0 = asset_master.get(nkey)
        if am0:
            rk = _normalize_name(am0.get("report_name", "") or "")
            if rk and rk in price_series:
                matched_ts = price_series[rk]
    # 4) 부분일치
    if matched_ts is None:
        for sk, sv in price_series.items():
            if nkey and (nkey in sk or sk in nkey):
                matched_ts = sv
                break
    if matched_ts is not None:
        matched_name = matched_ts["name"]

    if matched_ts is None:
        return {"name": key, "dates": [], "price_index": [], "bm_index": [],
                "bm_name": None, "period": period or "MAX", "as_of": None,
                "error": "종목 시계열 없음"}

    # 전체 날짜축 정렬
    all_dates = sorted(matched_ts["values"])

    # period 필터
    p = str(period or "MAX").upper()
    keep_map = {"1M": 21, "3M": 63, "1Y": 252}
    keep = keep_map.get(p)
    if keep and len(all_dates) > keep:
        all_dates = all_dates[-keep:]

    # ffill 가격 시계열
    price_filled = _ffill(all_dates, matched_ts["values"])

    # 시작=100 리베이스
    base_price = next((v for v in price_filled if v and v > 0), None)
    price_index: list[float | None] = []
    for v in price_filled:
        if base_price and v:
            price_index.append(round(v / base_price * 100.0, 4))
        else:
            price_index.append(None)

    # BM 시계열: asset_master에서 bm_name 조회
    bm_name: str | None = None
    am = asset_master.get(nkey)
    if am is None:
        for ak, av in asset_master.items():
            if nkey and (nkey in ak or ak in nkey):
                am = av
                break
    if am:
        bm_name = am.get("bm_name")
    if not bm_name:
        bm_name = DEFAULT_BM_NAME

    bm_ts = _resolve_bm_series(bm_name, bm_all_series)
    bm_index: list[float | None] = []
    resolved_bm_name = bm_name
    if bm_ts:
        resolved_bm_name = bm_ts["name"]
        bm_filled = _ffill(all_dates, bm_ts["values"])
        base_bm = next((v for v in bm_filled if v and v > 0), None)
        for v in bm_filled:
            if base_bm and v:
                bm_index.append(round(v / base_bm * 100.0, 4))
            else:
                bm_index.append(None)
    else:
        bm_index = [None] * len(all_dates)

    return {
        "name": matched_name,
        "dates": all_dates,
        "price_index": price_index,
        "bm_index": bm_index,
        "bm_name": resolved_bm_name,
        "period": p,
        "as_of": all_dates[-1] if all_dates else None,
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_pnl_summary() -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_summary
        return mock_pnl_summary()
    wb = _load_wb()
    if wb is None:
        # 데모용 mock 보유종목 (Excel 없을 때)
        return {
            "holdings": [
                {"name": "삼성전자",    "live_code": "005930", "qty": 10000, "pnl": 0, "pnl_pct": 0, "value": 0},
                {"name": "SK하이닉스", "live_code": "000660", "qty": 3000,  "pnl": 0, "pnl_pct": 0, "value": 0},
                {"name": "한화손해보험","live_code": "000370", "qty": 50000, "pnl": 0, "pnl_pct": 0, "value": 0},
            ],
            "total_pnl": 0, "total_value": 0,
            "transactions": [], "error": "Excel 파일 없음 (backend/data/)",
        }

    holdings_raw = _sheet_to_dicts(wb, "현재보유")
    sells_raw = _sheet_to_dicts(wb, "매도내역")
    asset_master = _build_asset_master(wb)
    wb.close()
    price_map, price_meta = _build_price_map()
    bm_series = _build_bm_series()
    _dates, price_series = _build_price_series()

    def col(d: dict, *cands: str) -> str | None:
        for k in d:
            kk = str(k).replace(" ", "")
            if any(c.replace(" ", "") in kk for c in cands):
                return k
        return None

    if not holdings_raw:
        return {"holdings": [], "total_pnl": 0, "total_value": 0,
                "transactions": [], "error": "현재보유 시트 비어있음"}

    h0 = holdings_raw[0]
    name_c = col(h0, "종목", "종 목")
    qty_c  = col(h0, "주수", "수량")
    cost_c = col(h0, "투자원금", "원금")
    unit_c = col(h0, "취득단가", "단가")
    date_c = col(h0, "매입일", "매수일", "날짜")
    bm_c   = col(h0, "BM명", "BM")

    name_to_code = _build_krx_name_to_code()
    try:
        from .price_service import get_quote as _get_quote
    except Exception:
        _get_quote = None

    holdings, total_cost, total_value = [], 0.0, 0.0
    live_quote_dates: list[str] = []
    for h in holdings_raw:
        name = str(h.get(name_c, "")).strip() if name_c else ""
        if not name:
            continue
        qty  = float(h.get(qty_c, 0) or 0) if qty_c else 0
        cost = float(h.get(cost_c, 0) or 0) if cost_c else 0
        unit = float(h.get(unit_c, 0) or 0) if unit_c else 0
        acq  = h.get(date_c) if date_c else None
        bm   = str(h.get(bm_c) or "") if bm_c else ""

        pinfo = _resolve_price(name, price_map, asset_master)
        cur_price = pinfo["price"] if pinfo else 0.0
        price_kind = pinfo["kind"] if pinfo else "N/A"
        price_date = pinfo.get("price_date") if pinfo else None
        usd_converted = bool(pinfo.get("usd_converted")) if pinfo else False
        price_native = pinfo.get("price_native") if pinfo else None

        # live_price: KRX 상장 ETF/주식에 한해 price_service 현재가를 추가 노출.
        # 단순 표시용이 아니라 평가금액/YTD/누적 손익 산정에도 우선 반영한다.
        live_price = None
        live_change_pct = None
        live_source = None
        live_code = _resolve_code(name, name_to_code, asset_master)
        live_currency = "KRW"
        if live_code and _get_quote is not None:
            try:
                lq = _get_quote(live_code)
                if lq and lq.get("price") is not None:
                    live_price = lq.get("price")
                    live_change_pct = lq.get("change_pct")
                    live_source = lq.get("source")
                    live_currency = "KRW"
                    if lq.get("as_of"):
                        live_quote_dates.append(str(lq.get("as_of"))[:10])
            except Exception:
                pass
        if live_price is None and usd_converted:
            foreign_ticker = _resolve_foreign_ticker(name, asset_master)
            if foreign_ticker:
                lq = _get_yahoo_quote_usd(foreign_ticker)
                if lq and lq.get("price") is not None:
                    live_price = lq.get("price")
                    live_change_pct = lq.get("change_pct")
                    live_source = lq.get("source")
                    live_code = str(lq.get("ticker") or foreign_ticker)
                    live_currency = "USD"
                    if lq.get("as_of"):
                        live_quote_dates.append(str(lq.get("as_of"))[:10])

        # BUG2 FIX: Use the single authoritative USDKRW rate from price_meta instead of
        # reverse-deriving FX from stale Excel prices (cur_price / price_native), which
        # would apply an outdated Excel FX to live USD prices. Fall back to derivation
        # only when price_meta["usdkrw"] is missing or zero.
        _meta_usdkrw = price_meta.get("usdkrw") if price_meta else None
        if usd_converted and _meta_usdkrw:
            valuation_fx = float(_meta_usdkrw)
        elif usd_converted and price_native:
            valuation_fx = cur_price / price_native  # graceful fallback
        else:
            valuation_fx = 1.0
        valuation_raw = live_price if live_price is not None else (
            price_native if usd_converted and price_native is not None else cur_price
        )
        valuation_price_krw = (valuation_raw or 0.0) * valuation_fx
        if (not cost) and unit and qty:
            # 해외주식(예: GRAB)은 현재보유 시트의 취득단가가 USD로 들어오고
            # 투자원금이 비어 있을 수 있다. 이 경우 단가×수량을 동일 FX로 KRW 환산해
            # 누적손익/수익률 분모로 사용한다.
            cost = unit * qty * valuation_fx if usd_converted else unit * qty
        # 기준가(펀드)는 좌수×기준가/1000 등 별도 환산이 필요할 수 있으나,
        # 1차 구현은 현재가×주수로 평가하고 price_kind를 함께 노출해 검증 가능케 한다.
        cur_value = valuation_price_krw * qty if (valuation_price_krw and qty) else 0.0
        pnl = cur_value - cost if cost else 0.0
        pnl_pct = (pnl / cost * 100) if cost else 0.0
        daily_pnl = None
        daily_pnl_pct = None
        ytd_pnl = None
        ytd_pnl_pct = None
        ytd_start = f"{dt.date.today().year}-01-01"
        acq_key = _to_date_str(acq)
        if acq_key and acq_key > ytd_start:
            ytd_start = acq_key
        ts = price_series.get(_normalize_name(pinfo["name"])) if pinfo else None
        if ts is None:
            # 직접 키 실패 시 보유명 기준 견고한 매칭(직접 → 부분일치)
            _nkn = _normalize_name(name)
            ts = price_series.get(_nkn)
            if ts is None:
                for _sk, _sv in price_series.items():
                    if _nkn and (_nkn in _sk or _sk in _nkn):
                        ts = _sv
                        break
        daily_pnl_source = None
        if ts:
            current_raw = valuation_raw
            fx = valuation_fx
            # ?? = Price_Raw ???? ? vs ? ?? ?? ??? ?.
            # ?? ??(??)? ???? forward-fill ? ???? ??? ? ??? ?? ?? ??? ??.
            _vals = ts["values"]
            _latest_date, last_raw, prev_raw = _latest_and_previous_values(_vals, price_date)
            if live_price is None and last_raw is not None:
                current_raw = last_raw
            if prev_raw and current_raw is not None:
                daily_pnl = (current_raw - prev_raw) * qty * fx
                daily_pnl_pct = (current_raw / prev_raw - 1.0) * 100
                daily_pnl_source = "price_raw"
            base_raw = _first_value_on_or_after(_vals, ytd_start)
            if base_raw and current_raw:
                ytd_pnl = (current_raw - base_raw) * qty * fx
                ytd_pnl_pct = (current_raw / base_raw - 1.0) * 100

        bm_master = asset_master.get(_normalize_name(name), {})
        adjusted_bm = bm or bm_master.get("bm_name") or ""
        bm_ts = _resolve_bm_series(adjusted_bm, bm_series)
        bm_return_pct = _bm_return_pct(bm_ts["values"], _to_date_str(acq)) if bm_ts else None
        excess_vs_bm_pct = round(pnl_pct - bm_return_pct, 2) if bm_return_pct is not None else None
        # daily_pnl: live_change_pct(??? ????%) ??, ??? Price_Raw ??? ??.
        # ?????? ?? ??? ????? ?? live ????? ?? ???? ????.
        final_daily_pnl = daily_pnl
        final_daily_pnl_pct = daily_pnl_pct
        final_daily_pnl_source = daily_pnl_source
        daily_value = None
        if live_change_pct is not None and live_price is not None and qty:
            live_value = cur_value
            r = live_change_pct / 100.0
            prev_value = live_value / (1.0 + r) if r > -0.999999 else live_value
            final_daily_pnl = round(live_value - prev_value, 0)
            final_daily_pnl_pct = round(live_change_pct, 2)
            final_daily_pnl_source = "live_quote"
            daily_value = round(live_value, 0)
        elif daily_pnl is not None:
            final_daily_pnl = round(daily_pnl, 0)
            final_daily_pnl_pct = round(daily_pnl_pct, 2) if daily_pnl_pct is not None else None
            daily_value = round(cur_value, 0)
        holdings.append({
            "name": name,
            "qty": qty,
            "cost": round(cost, 0),
            "unit_cost": round(unit, 2),
            "price": round(cur_price, 2),
            "price_kind": price_kind,
            "price_date": price_date,
            "usd_converted": usd_converted,
            "price_currency": "USD" if usd_converted else "KRW",
            "price_native": round(price_native, 4) if price_native is not None else None,
            "value": round(cur_value, 0),
            "pnl": round(pnl, 0),
            "pnl_pct": round(pnl_pct, 2),
            "daily_value": daily_value,
            "daily_pnl": final_daily_pnl,
            "daily_pnl_pct": final_daily_pnl_pct,
            "daily_pnl_source": final_daily_pnl_source,
            "ytd_pnl": round(ytd_pnl, 0) if ytd_pnl is not None else None,
            "ytd_pnl_pct": round(ytd_pnl_pct, 2) if ytd_pnl_pct is not None else None,
            "bm": adjusted_bm,
            "bm_return_pct": bm_return_pct,
            "excess_vs_bm_pct": excess_vs_bm_pct,
            "excess_pct": excess_vs_bm_pct,  # 프론트 계약용 별칭 (excess_vs_bm_pct 동일값)
            "contribution_pct": None,  # 전체 pnl 합산 후 아래서 채움
            "acq_date": str(acq)[:10] if acq else "",
            "matched": pinfo is not None,
            "live_price": live_price,
            "live_change_pct": live_change_pct,
            "live_code": live_code,
            "live_source": live_source,
            "live_currency": live_currency if live_price is not None else None,
        })
        total_cost += cost
        total_value += cur_value

    # BUG3 FIX: contribution_pct = each holding's share of TOTAL P&L (sums to ~100%).
    # Previous code used Σ|pnl| as denominator, which never summed to 100%.
    # When total_pnl is near zero (degenerate portfolio), fall back to Σ|pnl| to avoid
    # division instability.
    total_pnl_for_contrib = sum(h["pnl"] for h in holdings)
    denom = total_pnl_for_contrib if abs(total_pnl_for_contrib) > 1e-9 else (
        sum(abs(h["pnl"]) for h in holdings) or 1.0
    )
    for h in holdings:
        h["contribution_pct"] = round(h["pnl"] / denom * 100, 2)

    # total_daily_pnl: daily_pnl ?? (None ??)
    total_daily_pnl = sum(h["daily_pnl"] for h in holdings if h["daily_pnl"] is not None)
    total_daily_prev_value = sum(
        (h.get("daily_value") if h.get("daily_value") is not None else (h.get("value") or 0)) - (h.get("daily_pnl") or 0)
        for h in holdings
        if h.get("daily_pnl") is not None
    )
    total_daily_pnl_pct = (
        total_daily_pnl / total_daily_prev_value * 100
        if total_daily_prev_value
        else None
    )

    # Task B: header-sniffed transactions build
    sells_col_map, sells_fallback = _sniff_sells_cols(sells_raw)
    transactions = []
    for t in sells_raw[:30]:
        raw_name = t.get(sells_col_map["name"]) if sells_col_map["name"] else None
        if not raw_name:
            continue
        raw_date = t.get(sells_col_map["date"]) if sells_col_map["date"] else None
        raw_qty  = t.get(sells_col_map["qty"])  if sells_col_map["qty"]  else None
        raw_pr   = t.get(sells_col_map["pr"])   if sells_col_map["pr"]   else None
        raw_tr   = t.get(sells_col_map["tr"])   if sells_col_map["tr"]   else None
        transactions.append({
            "name": str(raw_name),
            "date": str(raw_date)[:10] if raw_date else "",
            "qty": str(raw_qty) if raw_qty is not None else "",
            "realized_pr": _pr_value_krw(raw_pr),
            "realized_tr": _tr_value_pct(raw_tr),
        })

    # realized_pnl_total: 매도내역 PR 실현손익 합산 (억원 단위 → 원 환산)
    realized_pnl_total = _calc_realized_pnl_total(sells_raw)

    total_pnl = total_value - total_cost
    return {
        "holdings": holdings,
        "total_cost": round(total_cost, 0),
        "total_value": round(total_value, 0),
        "total_pnl": round(total_pnl, 0),
        "total_pnl_pct": round((total_pnl / total_cost * 100) if total_cost else 0, 2),
        "total_daily_pnl": round(total_daily_pnl, 0),
        "total_daily_pnl_pct": round(total_daily_pnl_pct, 2) if total_daily_pnl_pct is not None else None,
        "realized_pnl_total": realized_pnl_total,
        "transactions": transactions,
        "unmatched": [h["name"] for h in holdings if not h["matched"]],
        "zero_price": [h["name"] for h in holdings if not h["price"]],
        "price_count": sum(1 for h in holdings if h["price"] > 0),
        "as_of": dt.datetime.now().isoformat(timespec="seconds"),
        # P0-6: 시세 기준일(Price_Raw 최신 일자)과 서버 조회 시각 분리
        "price_as_of": price_meta.get("price_as_of"),
        "live_price_as_of": max(live_quote_dates) if live_quote_dates else None,
        "fetched_at": dt.datetime.now().isoformat(timespec="seconds"),
        "usdkrw": price_meta.get("usdkrw"),
        "usdkrw_date": price_meta.get("usdkrw_date"),
        # Task B: warn when positional fallback was used for 매도내역 columns
        "sells_col_fallback": sells_fallback,
    }


# ---------------------------------------------------------------------------
# Task C: /api/pnl/trades — paginated + enriched realized trades
# ---------------------------------------------------------------------------

def get_pnl_trades(
    limit: int = 50,
    offset: int = 0,
    sort: str = "date",
    order: str = "desc",
) -> dict[str, Any]:
    if _mock_enabled():
        from .mock_portfolio import mock_pnl_trades
        return mock_pnl_trades(limit, offset, sort, order)
    """매도내역에서 실현 거래 목록을 반환 (페이지네이션 + 풍부한 컬럼).

    반환 shape:
      {
        trades: [{
            name,          # 종목명
            sell_date,     # 매도일 (YYYY-MM-DD or "")
            qty,           # 수량 (str, 원본 그대로)
            avg_buy_price, # 매입단가 (float or null)
            avg_sell_price,# 매도단가 (float or null)
            holding_days,  # 보유일수 (int or null) — 매입일 없으면 null
            pnl,           # 실현손익 PR (원, float or null)
            return_pct,    # 실현수익률 TR (%, float or null)
            cum_pnl,       # 날짜순 누적 실현손익 (원, float)
        }],
        total: 전체 행 수,
        limit: 요청 limit,
        offset: 요청 offset,
        sort: 정렬 기준,
        order: 정렬 방향,
        sells_col_fallback: bool,  # Task B: 위치 기반 폴백 사용 여부
        fetched_at: ISO 시각,
      }
    """
    wb = _load_wb()
    sells_raw = _sheet_to_dicts(wb, "매도내역") if wb else []
    if wb:
        wb.close()

    col_map, fallback_used = _sniff_sells_cols(sells_raw)

    # --- 전체 행 파싱 ---
    all_trades: list[dict[str, Any]] = []
    for t in sells_raw:
        raw_name = t.get(col_map["name"]) if col_map["name"] else None
        if not raw_name:
            continue

        raw_date      = t.get(col_map["date"])       if col_map["date"]       else None
        raw_qty       = t.get(col_map["qty"])        if col_map["qty"]        else None
        raw_buy_price = t.get(col_map["buy_price"])  if col_map["buy_price"]  else None
        raw_sell_price= t.get(col_map["sell_price"]) if col_map["sell_price"] else None
        raw_pr        = t.get(col_map["pr"])         if col_map["pr"]         else None
        raw_tr        = t.get(col_map["tr"])         if col_map["tr"]         else None

        sell_date = _to_date_str(raw_date) if raw_date else ""

        # qty: keep as string for display; also try numeric for holding_days
        qty_str = str(raw_qty) if raw_qty is not None else ""

        try:
            avg_buy_price: float | None = float(raw_buy_price) if raw_buy_price is not None else None
        except (TypeError, ValueError):
            avg_buy_price = None

        try:
            avg_sell_price: float | None = float(raw_sell_price) if raw_sell_price is not None else None
        except (TypeError, ValueError):
            avg_sell_price = None

        pnl = _pr_value_krw(raw_pr)
        return_pct = _tr_value_pct(raw_tr)

        # holding_days: requires both a buy date column and sell date
        # We look for a buy-date column separate from the sell-date column
        buy_date_col = _col_sniff(t, "매입일", "매수일", "취득일") if t else None
        holding_days: int | None = None
        if buy_date_col and sell_date:
            raw_buy_date = t.get(buy_date_col)
            buy_ds = _to_date_str(raw_buy_date) if raw_buy_date else None
            if buy_ds and sell_date:
                try:
                    bd = dt.date.fromisoformat(buy_ds)
                    sd = dt.date.fromisoformat(sell_date)
                    holding_days = (sd - bd).days
                    if holding_days < 0:
                        holding_days = None
                except (ValueError, TypeError):
                    holding_days = None

        all_trades.append({
            "name": str(raw_name),
            "sell_date": sell_date,
            "qty": qty_str,
            "avg_buy_price": avg_buy_price,
            "avg_sell_price": avg_sell_price,
            "holding_days": holding_days,
            "pnl": pnl,
            "return_pct": return_pct,
            "cum_pnl": 0.0,  # filled below after sort
        })

    # --- 날짜순 정렬 후 cum_pnl 계산 (정렬 방향과 무관하게 날짜 오름차순으로) ---
    dated = sorted(all_trades, key=lambda x: x["sell_date"] or "")
    cum = 0.0
    for tr in dated:
        if tr["pnl"] is not None:
            cum += tr["pnl"]
        tr["cum_pnl"] = round(cum, 0)

    # --- 요청한 sort/order 적용 ---
    valid_sort = {"date", "pnl", "return_pct", "holding_days", "name"}
    sort_key = sort if sort in valid_sort else "date"
    reverse = (order.lower() != "asc")

    def _sort_val(tr: dict) -> Any:
        v = tr.get({"date": "sell_date"}.get(sort_key, sort_key))
        if v is None:
            # None values sink to the end regardless of order
            return ("" if reverse else "\xff") if isinstance(v, str) else (float("inf") if reverse else float("-inf"))
        return v

    # stable sort: Nones to end
    none_trades = [tr for tr in dated if _sort_val(tr) in (None,)]
    val_trades  = [tr for tr in dated if _sort_val(tr) not in (None,)]

    def _safe_sort_val(tr: dict) -> Any:
        v = tr["sell_date"] if sort_key == "date" else tr.get(sort_key)
        if v is None:
            return ""
        return v

    val_trades.sort(key=_safe_sort_val, reverse=reverse)
    sorted_trades = val_trades + none_trades

    total = len(sorted_trades)
    page = sorted_trades[offset: offset + limit]

    return {
        "trades": page,
        "total": total,
        "limit": limit,
        "offset": offset,
        "sort": sort_key,
        "order": order,
        "sells_col_fallback": fallback_used,
        "fetched_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
